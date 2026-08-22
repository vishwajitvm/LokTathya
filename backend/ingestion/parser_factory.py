import json
import csv
import io
from typing import Dict, Any, List, Optional
try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

from ingestion.html_normalizer import HTMLNormalizer
from core.format_detector import FormatDetector

class ParserFactory:
    """
    Priority 1: Multi-format parsing factory for HTML, PDF, CSV, XLSX, JSON.
    Conforms to a strict, standardized output model.
    """

    @staticmethod
    def get_parser(detected_format: str):
        if detected_format == "HTML":
            return HTMLParser()
        elif detected_format == "PDF":
            return PDFParser()
        elif detected_format == "CSV":
            return CSVParser()
        elif detected_format in ["XLSX", "XLSM", "XLS"]:
            return XLSXParser()
        elif detected_format == "JSON":
            return JSONParser()
        elif detected_format in ["GEOJSON", "KML", "KMZ", "SHAPEFILE_ARCHIVE", "GIS"]:
            return GISParser()
        elif detected_format == "XML":
            return XMLParser()
        elif detected_format == "ZIP":
            return GISParser()
        else:
            return PlainTextParser()

class BaseParser:
    def parse(self, content: bytes, meta: Dict[str, Any]) -> Dict[str, Any]:
        raise NotImplementedError

    def _build_result(
        self,
        meta: Dict[str, Any],
        structured_content: Any,
        text_content: Optional[str] = None,
        parser_name: str = "Base",
        parser_version: str = "1.0.0",
        status: str = "SUCCESS"
    ) -> Dict[str, Any]:
        return {
            "document_id": meta.get("document_id"),
            "version_id": meta.get("version_id"),
            "source_id": meta.get("source_id"),
            "parser_name": parser_name,
            "parser_version": parser_version,
            "status": status,
            "metadata": meta.get("custom_metadata", {}),
            "structured_content": structured_content,
            "text_content": text_content,
            "extraction_location": meta.get("storage_path", "unknown")
        }

class HTMLParser(BaseParser):
    def parse(self, content: bytes, meta: Dict[str, Any]) -> Dict[str, Any]:
        html_str = content.decode('utf-8', errors='ignore')
        normalizer = HTMLNormalizer(html_str)
        norm_res = normalizer.normalize()
        tables = normalizer.extract_tables()
        
        # Link extraction
        from ingestion.link_extractor import DocumentLinkExtractor
        extractor = DocumentLinkExtractor(html_str, meta.get("base_url", "https://gov.in"))
        links = extractor.extract_links()

        structured = {
            "tables": tables,
            "links": links,
            "title": norm_res.get("title")
        }

        return self._build_result(
            meta=meta,
            structured_content=structured,
            text_content=norm_res["text"],
            parser_name="HTMLParser"
        )

class PDFParser(BaseParser):
    def parse(self, content: bytes, meta: Dict[str, Any]) -> Dict[str, Any]:
        text_runs = []
        page_count = 0
        if PdfReader:
            try:
                reader = PdfReader(io.BytesIO(content))
                page_count = len(reader.pages)
                if page_count > 500:
                    return self._build_result(meta=meta, structured_content={}, status="ERROR: PDF page count exceeds maximum limit (500 pages)", parser_name="PDFParser")
                for page in reader.pages:
                    txt = page.extract_text() or ""
                    text_runs.append(txt)
            except Exception as e:
                return self._build_result(meta=meta, structured_content={}, status=f"ERROR: {str(e)}", parser_name="PDFParser")
        else:
            text_runs.append("pypdf missing in runtime environment")

        full_text = "\n--- PAGE BREAK ---\n".join(text_runs)
        
        return self._build_result(
            meta=meta,
            structured_content={"page_count": page_count},
            text_content=full_text,
            parser_name="PDFParser"
        )

class CSVParser(BaseParser):
    def parse(self, content: bytes, meta: Dict[str, Any]) -> Dict[str, Any]:
        if len(content) > 50 * 1024 * 1024:
            return self._build_result(meta=meta, structured_content={}, status="ERROR: File size exceeds maximum limit (50MB)", parser_name="CSVParser")

        # Detect encoding
        encoding = FormatDetector._detect_csv_encoding(content)
        csv_str = content.decode(encoding, errors='ignore')
        
        # Detect delimiter
        dialect = None
        try:
            dialect = csv.Sniffer().sniff(csv_str[:4096])
        except Exception:
            pass
            
        delimiter = dialect.delimiter if dialect else ','
        
        rows = []
        reader = csv.reader(io.StringIO(csv_str), delimiter=delimiter)
        for idx, row in enumerate(reader):
            if idx >= 100000:
                break
            rows.append(row)
            
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        structured = {
            "headers": headers,
            "rows": data_rows,
            "delimiter": delimiter,
            "encoding": encoding
        }
        
        return self._build_result(
            meta=meta,
            structured_content=structured,
            text_content=csv_str,
            parser_name="CSVParser"
        )

class XLSXParser(BaseParser):
    def parse(self, content: bytes, meta: Dict[str, Any]) -> Dict[str, Any]:
        if len(content) > 50 * 1024 * 1024:
            return self._build_result(meta=meta, structured_content={}, status="ERROR: File size exceeds maximum limit (50MB)", parser_name="XLSXParser")

        sheets = {}
        if openpyxl:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                total_rows = 0
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    sheet_rows = []
                    for row in ws.iter_rows(values_only=True):
                        total_rows += 1
                        if total_rows >= 100000:
                            break
                        # Convert cells to string or primitives
                        sheet_rows.append([str(cell) if cell is not None else "" for cell in row])
                    sheets[sheet_name] = {
                        "headers": sheet_rows[0] if sheet_rows else [],
                        "rows": sheet_rows[1:] if len(sheet_rows) > 1 else []
                    }
                    if total_rows >= 100000:
                        break
            except Exception as e:
                return self._build_result(meta=meta, structured_content={}, status=f"ERROR: {str(e)}", parser_name="XLSXParser")
        else:
            sheets["error"] = "openpyxl missing in environment"

        return self._build_result(
            meta=meta,
            structured_content={"sheets": sheets},
            text_content=json.dumps(sheets, indent=2),
            parser_name="XLSXParser"
        )

class JSONParser(BaseParser):
    def parse(self, content: bytes, meta: Dict[str, Any]) -> Dict[str, Any]:
        try:
            data = json.loads(content.decode('utf-8', errors='ignore'))
        except Exception as e:
            return self._build_result(meta=meta, structured_content={}, status=f"ERROR: {str(e)}", parser_name="JSONParser")

        return self._build_result(
            meta=meta,
            structured_content=data,
            text_content=json.dumps(data, indent=2),
            parser_name="JSONParser"
        )

class GISParser(BaseParser):
    def parse(self, content: bytes, meta: Dict[str, Any]) -> Dict[str, Any]:
        import zipfile
        import xml.etree.ElementTree as ET
        detected_format = meta.get("detected_format", "GEOJSON")
        structured = {}
        text = ""
        
        try:
            if detected_format == "GEOJSON" or content.startswith(b"{"):
                data = json.loads(content.decode('utf-8', errors='ignore'))
                structured = {
                    "type": data.get("type", "FeatureCollection"),
                    "features_count": len(data.get("features", [])),
                    "crs": data.get("crs", {"type": "name", "properties": {"name": "EPSG:4326"}})
                }
                text = f"GeoJSON FeatureCollection with {structured['features_count']} features."
                
            elif detected_format == "KML" or b"<kml" in content.lower():
                if b"<!ENTITY" in content or b"<!DOCTYPE" in content:
                    return self._build_result(meta=meta, structured_content={}, status="ERROR: XML entity declarations or DOCTYPES are forbidden for security reasons.", parser_name="GISParser")
                root = ET.fromstring(content)
                placemarks = root.findall(".//Placemark")
                structured = {
                    "placemarks_count": len(placemarks),
                    "document_name": "KML Document"
                }
                text = f"KML Document containing {structured['placemarks_count']} placemarks."
                
            elif detected_format == "KMZ":
                with zipfile.ZipFile(io.BytesIO(content)) as z:
                    total_size = sum(info.file_size for info in z.infolist())
                    if total_size > 200 * 1024 * 1024:
                        return self._build_result(meta=meta, structured_content={}, status="ERROR: Decompressed archive size exceeds maximum limit (200MB)", parser_name="GISParser")
                    if len(z.infolist()) > 1000:
                        return self._build_result(meta=meta, structured_content={}, status="ERROR: Archive contains too many files (max 1000)", parser_name="GISParser")

                    for name in z.namelist():
                        if ".." in name or name.startswith("/") or name.startswith("\\") or ":" in name:
                            return self._build_result(meta=meta, structured_content={}, status="ERROR: Malicious path traversal inside archive (Zip Slip Prevention)", parser_name="GISParser")
                    
                    kml_names = [n for n in z.namelist() if n.endswith('.kml')]
                    if kml_names:
                        kml_content = z.read(kml_names[0])
                        if b"<!ENTITY" in kml_content or b"<!DOCTYPE" in kml_content:
                            return self._build_result(meta=meta, structured_content={}, status="ERROR: XML entity declarations or DOCTYPES are forbidden for security reasons.", parser_name="GISParser")
                        root = ET.fromstring(kml_content)
                        placemarks = root.findall(".//Placemark")
                        structured = {
                            "placemarks_count": len(placemarks),
                            "document_name": kml_names[0]
                        }
                        text = f"KMZ Document containing KML: {kml_names[0]} with {structured['placemarks_count']} placemarks."
                    else:
                        structured = {"error": "No KML file found in KMZ archive"}
                        
            elif detected_format in ["SHAPEFILE_ARCHIVE", "ZIP"]:
                with zipfile.ZipFile(io.BytesIO(content)) as z:
                    total_size = sum(info.file_size for info in z.infolist())
                    if total_size > 200 * 1024 * 1024:
                        return self._build_result(meta=meta, structured_content={}, status="ERROR: Decompressed archive size exceeds maximum limit (200MB)", parser_name="GISParser")
                    if len(z.infolist()) > 1000:
                        return self._build_result(meta=meta, structured_content={}, status="ERROR: Archive contains too many files (max 1000)", parser_name="GISParser")

                    names = z.namelist()
                    for name in names:
                        if ".." in name or name.startswith("/") or name.startswith("\\") or ":" in name:
                            return self._build_result(meta=meta, structured_content={}, status="ERROR: Malicious path traversal inside archive (Zip Slip Prevention)", parser_name="GISParser")
                    
                    structured = {
                        "files": names,
                        "is_shapefile": any(n.endswith('.shp') for n in names)
                    }
                    text = f"ZIP Archive containing {len(names)} files. Is Shapefile: {structured['is_shapefile']}"
        except Exception as e:
            return self._build_result(meta=meta, structured_content={}, status=f"ERROR: {str(e)}", parser_name="GISParser")

        return self._build_result(
            meta=meta,
            structured_content=structured,
            text_content=text,
            parser_name="GISParser"
        )

class XMLParser(BaseParser):
    def parse(self, content: bytes, meta: Dict[str, Any]) -> Dict[str, Any]:
        import xml.etree.ElementTree as ET
        if b"<!ENTITY" in content or b"<!DOCTYPE" in content:
            return self._build_result(meta=meta, structured_content={}, status="ERROR: XML entity declarations or DOCTYPES are forbidden for security reasons.", parser_name="XMLParser")
        try:
            root = ET.fromstring(content)
            structured = {
                "tag": root.tag,
                "attrib": dict(root.attrib),
                "children_count": len(list(root))
            }
            text = f"XML document root: {root.tag}"
        except Exception as e:
            return self._build_result(meta=meta, structured_content={}, status=f"ERROR: {str(e)}", parser_name="XMLParser")

        return self._build_result(
            meta=meta,
            structured_content=structured,
            text_content=text,
            parser_name="XMLParser"
        )

class PlainTextParser(BaseParser):
    def parse(self, content: bytes, meta: Dict[str, Any]) -> Dict[str, Any]:
        text = content.decode('utf-8', errors='ignore')
        return self._build_result(
            meta=meta,
            structured_content={"lines_count": len(text.splitlines())},
            text_content=text,
            parser_name="PlainTextParser"
        )
